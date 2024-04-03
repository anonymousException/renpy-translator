import _thread
import os.path
import time
import traceback

import openpyxl
from PySide6.QtCore import QCoreApplication, Qt, QSortFilterProxyModel, QAbstractTableModel, QModelIndex, QThread, \
    Signal
from PySide6.QtGui import QIntValidator, QIcon
from PySide6.QtWidgets import QDialog, QTableWidget, QTableView, QHeaderView, QTableWidgetItem, QFileDialog, \
    QInputDialog, QMessageBox, QStyledItemDelegate, QPushButton, QLineEdit
from openpyxl.workbook import Workbook

from local_glossary import Ui_LocalGlossaryDialog
from my_log import log_print, log_path


class LineEditDelegate(QStyledItemDelegate):

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        return editor

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.EditRole)
        editor.setText(str(index.data()))


def save_to_xlsx_file(filename, wb):
    wb.save(filename)


class ExcelModel(QAbstractTableModel):
    def __init__(self, data, path, wb, rows_per_page=100):
        super().__init__()
        self.is_need_save = False
        if data is None:
            self.init = False
            return
        self.init = True
        self._data = data
        self.path = path
        self.wb = wb
        self.rows_per_page = rows_per_page
        self.current_page = 0
        self.max_row = self._data.max_row
        self.max_column = self._data.max_column
        self.header_titles = [QCoreApplication.translate('LocalGlossaryDialog', 'Row', None)] + [
            self._data.cell(row=1, column=col).value for col in
            range(1, self.max_column + 1)]
        self.dataChanged.connect(self.handle_data_changed)

    def handle_data_changed(self, top_left, bottom_right):
        self.is_need_save = True

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags

        if index.column() == 0:
            return super().flags(index)
        else:
            return super().flags(index) | Qt.ItemIsEditable

    def setData(self, index, value, role=Qt.EditRole):
        if not index.isValid() or role != Qt.EditRole:
            return False
        if index.column() == 0:
            return False
        row = self.current_page * self.rows_per_page + index.row() + 2
        column = index.column()
        self._data.cell(row=row, column=column).value = value
        self.dataChanged.emit(index, index)
        return True

    def save_data(self):
        if os.path.isfile(self.path):
            self.wb.save(self.path)

    def rowCount(self, parent=QModelIndex()):
        return min(self.rows_per_page, self.max_row - self.current_page * self.rows_per_page - 1)

    def columnCount(self, parent=QModelIndex()):
        return self.max_column + 1

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or role != Qt.DisplayRole:
            return None

        if index.column() == 0:
            value = self.current_page * self.rows_per_page + index.row() + 1
            if value is None:
                value = ''
            return str(value)
        else:
            row = self.current_page * self.rows_per_page + index.row() + 2
            column = index.column()
            if row <= self.max_row:
                cell = self._data.cell(row=row, column=column)
                if cell.value is None:
                    cell.value = ''
                return str(cell.value)

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self.header_titles[section]
        else:
            return str(self.current_page * self.rows_per_page + section + 1)

    def setPage(self, page):
        self.beginResetModel()
        self.current_page = page
        self.endResetModel()

    def resetPageRows(self, rows):
        self.beginResetModel()
        self.rows_per_page = rows
        self.endResetModel()


class MyTableView(QTableView):
    def __init__(self, parent=None):
        super(MyTableView, self).__init__(parent)
        self.model = ExcelModel(None, None, None)
        # self.setModel(self.model)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.header_titles = ([QCoreApplication.translate('LocalGlossaryDialog', 'Row', None),
                               QCoreApplication.translate('LocalGlossaryDialog', 'Original', None),
                               QCoreApplication.translate('LocalGlossaryDialog', 'Replace', None)])
        self.file = None
        self.row = 0
        self.rows_to_hide = []
        self.verticalHeader().setVisible(False)
        lineEditDelegate = LineEditDelegate()
        self.setItemDelegate(lineEditDelegate)


class MyLocalGlossaryForm(QDialog, Ui_LocalGlossaryDialog):
    def __init__(self, parent=None):
        super(MyLocalGlossaryForm, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('main.ico'))
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)
        self.tableView = MyTableView()
        self.tableView.selectFileText = self.selectFileText
        self.verticalLayout.addWidget(self.tableView)
        self.selectFileBtn.clicked.connect(self.select_file)
        self.selectFileText.textChanged.connect(self.on_text_changed)
        self.data = None
        self.confirmButton.clicked.connect(self.on_confirm_clicked)
        self.create_file_param = None
        self.gotoPageLineEdit.setValidator(QIntValidator(1, 0, self))
        self.pageRowsLineEdit.setValidator(QIntValidator(1, 9999, self))
        _thread.start_new_thread(self.update, ())

        def on_previous_clicked():
            if not self.tableView.model.init:
                return
            new_page = max(self.tableView.model.current_page - 1, 0)
            self.tableView.model.setPage(new_page)
            self.curPageLabel.setText(str(self.tableView.model.current_page + 1) + '/' + str(
                int(self.tableView.model.max_row / self.tableView.model.rows_per_page) + 1))

        def on_next_clicked():
            if not self.tableView.model.init:
                return
            new_page = min(self.tableView.model.current_page + 1,
                           (self.tableView.model.max_row - 1) // self.tableView.model.rows_per_page)
            self.tableView.model.setPage(new_page)
            self.curPageLabel.setText(str(self.tableView.model.current_page + 1) + '/' + str(
                int(self.tableView.model.max_row / self.tableView.model.rows_per_page) + 1))

        self.prevButton.clicked.connect(on_previous_clicked)
        self.nextButton.clicked.connect(on_next_clicked)
        self.gotoPageButton.clicked.connect(self.on_goto_clicked)
        self.maxPageRowsButton.clicked.connect(self.on_max_rows_clicked)
        self.gotoPageLineEdit.returnPressed.connect(self.on_goto_clicked)
        self.pageRowsLineEdit.returnPressed.connect(self.on_max_rows_clicked)

    def on_goto_clicked(self):
        if len(self.gotoPageLineEdit.text()) == 0:
            return
        page = int(self.gotoPageLineEdit.text()) - 1
        if page < 0:
            page = 0
        max_page = int(self.tableView.model.max_row / self.tableView.model.rows_per_page)
        if page > max_page:
            page = max_page
            self.gotoPageLineEdit.setText(str(page))
        self.tableView.model.setPage(page)
        self.curPageLabel.setText(str(self.tableView.model.current_page + 1) + '/' + str(
            int(self.tableView.model.max_row / self.tableView.model.rows_per_page) + 1))

    def on_max_rows_clicked(self):
        if len(self.pageRowsLineEdit.text()) == 0:
            return
        max_rows = int(self.pageRowsLineEdit.text())
        self.tableView.model.resetPageRows(max_rows)
        self.gotoPageLineEdit.setText(str(0))
        self.on_goto_clicked()
        self.curPageLabel.setText(str(self.tableView.model.current_page + 1) + '/' + str(
            int(self.tableView.model.max_row / self.tableView.model.rows_per_page) + 1))

    def on_text_changed(self):
        file = self.selectFileText.toPlainText().replace('file:///', '')
        if len(file) == 0:
            return
        if os.path.isfile(file) and file.endswith('.xlsx'):
            self.load_from_xlsx(file)
            self.tableView.file = file

    def select_file(self):
        file, filetype = QFileDialog.getOpenFileName(self,
                                                     QCoreApplication.translate('LocalGlossaryDialog',
                                                                                'select the file you want to import',
                                                                                None),
                                                     '',
                                                     "Xlsx Files (*.xlsx)")
        if os.path.isfile(file):
            self.tableView.file = file
            self.selectFileText.setText(file)

    def load_from_xlsx(self, file):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        self.tableView.model = ExcelModel(ws, file, wb)
        self.tableView.setModel(self.tableView.model)
        self.curPageLabel.setText(str(self.tableView.model.current_page + 1) + '/' + str(
            int(self.tableView.model.max_row / self.tableView.model.rows_per_page) + 1))
        self.gotoPageLineEdit.setValidator(QIntValidator(1, self.tableView.model.max_row, self))
        self.on_max_rows_clicked()
        return

    def get_data(self):
        dic = dict()
        model = self.tableView.model
        rows = model.rowCount()
        cols = model.columnCount()
        if cols < 2:
            return None
        for r in range(rows):
            index1 = model.index(r, 1)
            index2 = model.index(r, 2)
            value1 = model.data(index1, Qt.DisplayRole)
            value2 = model.data(index2, Qt.DisplayRole)
            if value1 is None or value2 is None:
                continue
            dic[value1] = value2
        return dic

    def on_confirm_clicked(self):
        self.data = self.get_data()
        self.close()

    def update(self):
        thread = self.UpdateThread()
        thread.update_date.connect(self.update_progress)
        while True:
            thread.start()
            time.sleep(0.1)

    def update_progress(self):
        try:
            if self.tableView.model.is_need_save:
                self.tableView.model.save_data()
                self.tableView.model.is_need_save = False
        except Exception:
            msg = traceback.format_exc()
            log_print(msg)

    class UpdateThread(QThread):
        update_date = Signal()

        def __init__(self):
            super().__init__()

        def __del__(self):
            self.wait()

        def run(self):
            self.update_date.emit()
