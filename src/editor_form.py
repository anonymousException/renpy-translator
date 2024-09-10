import _thread
import json
import multiprocessing

import os.path
import io
import subprocess
import sys
import threading
import time
import traceback
import webbrowser
import pyperclip
from PySide6 import QtCore
from PySide6.QtCore import Qt, QDir, QModelIndex, QSortFilterProxyModel, Signal, QThread, QCoreApplication
from PySide6.QtGui import QStandardItemModel, QStandardItem, QIcon, QColor, QTextCursor, QKeySequence, QIntValidator
from PySide6.QtWidgets import QDialog, QHeaderView, QTableView, QMenu, QListView, QFileDialog, QTreeView, \
    QFileSystemModel, QStyle, QMessageBox, QButtonGroup, QInputDialog, QVBoxLayout, QLabel, QTextEdit, QPushButton, \
    QCheckBox, QLineEdit, QAbstractItemView
from openpyxl.workbook import Workbook

from editor import Ui_EditorDialog
from my_log import log_print, log_path
from renpy_translate import init_client, TranslateToList, engineDic, language_header, get_rpy_info, get_translated, \
    rpy_info_dic, web_brower_export_name, get_translated_dic
from custom_engine_form import targetDic, sourceDic
from engine_form import MyEngineForm
from local_glossary_form import MyLocalGlossaryForm
from export_setting_form import MyExportSettingForm
from html_util import write_html_with_strings
from import_html_form import MyImportHtmlForm
import html_util
from translated_form import MyTranslatedForm
from string_tool import *
from html_util import open_directory_and_select_file


translated_thread = None
translated_dic = None
is_need_save = False
rpy_lock = threading.Lock()

editor_main_trhead_lock = threading.Lock()
editor_main_thread_tasks = []

class CustomSortProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super(CustomSortProxyModel, self).__init__(parent)
        self.filter_pattern = '.rpy'

    def lessThan(self, left, right):
        if left.column() > 0 and right.column() > 0:
            leftValue = 0
            rightValue = 0
            if len(left.data()) > 0:
                leftValue = float(left.data().strip('%'))
            if len(right.data()) > 0:
                rightValue = float(right.data().strip('%'))
            return leftValue < rightValue
        else:
            return super().lessThan(left, right)

    def filterAcceptsRow(self, source_row, source_parent):
        index0 = self.sourceModel().index(source_row, 0, source_parent)
        file_name = self.sourceModel().fileName(index0)
        if self.sourceModel().isDir(index0) or file_name.endswith(self.filter_pattern):
            return True
        else:
            return False


class FileSystemModel(QFileSystemModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_open_filter = True

    def columnCount(self, parent=QModelIndex()):
        return super().columnCount(parent) + 2

    def data(self, index, role=Qt.DisplayRole):
        global rpy_info_dic
        if role == Qt.DisplayRole and index.column() == 4:
            path = index.sibling(index.row(), 0).data()
            full_path = self.filePath(index)
            # info = self.fileInfo(self.index(index.row(), 0, index.parent()))
            # full_path = info.absoluteFilePath()
            if os.path.isfile(full_path):
                if full_path in rpy_info_dic.keys():
                    ret, unmatch_cnt, p = rpy_info_dic[full_path]
                else:
                    return ''
                    ret, unmatch_cnt, p = get_rpy_info(full_path)
                rpy_info_dic[full_path] = ret, unmatch_cnt, p
                return str(len(ret))
            else:
                return ''
        elif role == Qt.DisplayRole and index.column() == 5:
            path = index.sibling(index.row(), 0).data()
            full_path = self.filePath(index)
            # info = self.fileInfo(self.index(index.row(), 0, index.parent()))
            # full_path = info.absoluteFilePath()
            if os.path.isfile(full_path):
                if full_path in rpy_info_dic.keys():
                    ret, unmatch_cnt, p = rpy_info_dic[full_path]
                else:
                    return ''
                    ret, unmatch_cnt, p = get_rpy_info(full_path)
                rpy_info_dic[full_path] = ret, unmatch_cnt, p
                if len(ret) != 0:
                    percentage = unmatch_cnt / len(ret) * 100
                else:
                    percentage = 0
                return f'{percentage:.2f}%'
            else:
                return ''
        else:
            return super().data(index, role)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if (orientation == Qt.Horizontal and
                role == Qt.DisplayRole and
                section == 0):
            return QCoreApplication.translate('EditorDialog', 'Path', None)
        if (orientation == Qt.Horizontal and
                role == Qt.DisplayRole and
                section == self.columnCount() - 2):
            return QCoreApplication.translate('EditorDialog', 'Units', None)
        if (orientation == Qt.Horizontal and
                role == Qt.DisplayRole and
                section == self.columnCount() - 1):
            return QCoreApplication.translate('EditorDialog', 'Translated', None)
        return super().headerData(section, orientation, role)


class MyTreeView(QTreeView):

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.setColumnWidth(0, self.width() * 0.5)

    def __init__(self, parent=None):
        super(MyTreeView, self).__init__(parent)
        self.setWindowTitle("Dir View")
        self.model = FileSystemModel()
        self.model.setRootPath('/')
        self.proxy_model = CustomSortProxyModel()
        self.proxy_model.setSourceModel(self.model)
        self.setModel(self.proxy_model)
        self.setRootIndex(self.proxy_model.mapFromSource(self.model.index('/')))
        for i in [1, 2, 3]:
            self.setColumnHidden(i, True)
        self.header().sectionClicked.connect(self.handle_header_clicked)
        self.setSortingEnabled(True)
        self.asc = True
        self.selectionModel().selectionChanged.connect(self.on_selection_changed)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.openMenu)
        self.export_path = None
        self.export_html = False
        self.import_html = False
        self.is_replace_special_symbols = True
        self.units_min = 0
        self.units_max = sys.maxsize
        self.translated_min = 0.0
        self.translated_max = 100.0
        self.is_fresh = False
        # self.hide()

    def openMenu(self, position):
        indexes = self.selectedIndexes()
        if len(indexes) > 0:
            menu = QMenu()
            selectTableView = MySelectTableView(self.selectTableView)
            index = self.currentIndex()
            source_index = self.proxy_model.mapToSource(index)
            file_info = QtCore.QFileInfo(self.model.filePath(source_index))
            full_path = file_info.absoluteFilePath()
            full_path = full_path.replace('\\', '/')
            menu.addAction(QCoreApplication.translate('EditorDialog', 'Export to xlsx file', None),
                           lambda: selectTableView.export_to_xlsx(full_path, None, self.model.is_open_filter, self))
            menu.addAction(QCoreApplication.translate('EditorDialog', 'Export to html file', None),
                           lambda: selectTableView.export_to_html(full_path, None, self.model.is_open_filter, self))
            menu.addAction(
                QCoreApplication.translate('EditorDialog', 'Import html and relative translated contents', None),
                lambda: selectTableView.import_from_html(full_path, None, self.model.is_open_filter, self))

            menu.exec(self.viewport().mapToGlobal(position))

    def refresh_table_view(self, full_path):
        global rpy_info_dic
        if os.path.isfile(full_path):
            self.tableView.model.clear()
            self.tableView.searched.clear()
            self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.tableView.model.setHorizontalHeaderLabels([QCoreApplication.translate('EditorDialog', 'line', None),
                                                            QCoreApplication.translate('EditorDialog', 'refer', None),
                                                            QCoreApplication.translate('EditorDialog', 'Original',
                                                                                       None),
                                                            QCoreApplication.translate('EditorDialog', 'Current', None),
                                                            QCoreApplication.translate('EditorDialog', 'Translated',
                                                                                       None)])
            if full_path in rpy_info_dic.keys():
                ret, unmatch_cnt, p = rpy_info_dic[full_path]
            else:
                ret, unmatch_cnt, p = get_rpy_info(full_path)
                rpy_info_dic[full_path] = ret, unmatch_cnt, p
            for i in ret:
                row = [
                    QStandardItem(str(i['line'])),
                    QStandardItem(i['refer']),
                    QStandardItem(i['original']),
                    QStandardItem(i['current']),
                    QStandardItem(''),
                ]
                for item in row:
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setToolTip(item.text())
                row[0].setData(i, Qt.UserRole)
                row[0].setEditable(False)
                row[1].setEditable(False)
                row[2].setEditable(False)
                self.tableView.model.appendRow(row)
            self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            self.tableView.file = full_path
            self.tableView.index = self.currentIndex()
            self.tableView.editorForm.on_checkbox_state_changed()

    def on_selection_changed(self, selected, deselected):
        selected_indexes = selected.indexes()
        if len(selected_indexes) > 0:
            index = self.currentIndex()
            source_index = self.proxy_model.mapToSource(index)
            file_info = QtCore.QFileInfo(self.model.filePath(source_index))
            full_path = file_info.absoluteFilePath()
            full_path = full_path.replace('\\', '/')
            self.refresh_table_view(full_path)

    def handle_header_clicked(self, index):
        if self.asc:
            self.sortByColumn(index, Qt.AscendingOrder)
            self.asc = False
        else:
            self.sortByColumn(index, Qt.DescendingOrder)
            self.asc = True


class MySelectTableView(QTableView):
    def __init__(self, parent=None):
        super(MySelectTableView, self).__init__(parent)
        self.model = QStandardItemModel()
        self.setModel(self.model)
        self.model.setHorizontalHeaderLabels([QCoreApplication.translate('EditorDialog', 'Path', None),
                                              QCoreApplication.translate('EditorDialog', 'Units', None),
                                              QCoreApplication.translate('EditorDialog', 'Translated', None)])
        self.verticalHeader().setVisible(False)
        self.setSelectionBehavior(QTableView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setSortingEnabled(True)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def load_data(self, index):
        global rpy_info_dic
        if index.isValid():
            select_one = index.data()
            if self.treeView.model.is_open_filter:
                self.treeView.proxy_model.filter_pattern = '.rpy'
            else:
                self.treeView.proxy_model.filter_pattern = ''
            if os.path.isdir(select_one):
                self.treeView.tableView.editorForm.setDisabled(True)
                rpy_info_dic.clear()
                self.tableView.model.clear()
                self.tableView.searched.clear()
                t = getRpyInfoThread(p=select_one, is_open_filter=self.treeView.model.is_open_filter)
                t.start()
                self.treeView.tableView.editorForm.setDisabled(True)
                self.treeView.show()
            else:
                if (os.path.isfile(select_one)):
                    rpy_info_dic.clear()
                    self.tableView.model.clear()
                    self.tableView.searched.clear()
                    self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                    self.tableView.model.setHorizontalHeaderLabels(
                        [QCoreApplication.translate('EditorDialog', 'line', None),
                         QCoreApplication.translate('EditorDialog', 'refer', None),
                         QCoreApplication.translate('EditorDialog', 'Original', None),
                         QCoreApplication.translate('EditorDialog', 'Current', None),
                         QCoreApplication.translate('EditorDialog', 'Translated', None)])

                    ret, unmatch_cnt, p = get_rpy_info(select_one)
                    rpy_info_dic[select_one] = ret, unmatch_cnt, p
                    for i in ret:
                        row = [
                            QStandardItem(str(i['line'])),
                            QStandardItem(i['refer']),
                            QStandardItem(i['original']),
                            QStandardItem(i['current']),
                            QStandardItem(''),
                        ]
                        for item in row:
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setToolTip(item.text())
                        row[0].setData(i, Qt.UserRole)
                        row[0].setEditable(False)
                        row[1].setEditable(False)
                        row[2].setEditable(False)
                        self.tableView.model.appendRow(row)
                    self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
                    self.tableView.file = select_one
                    self.tableView.index = index
                    self.tableView.editorForm.on_checkbox_state_changed()
                self.treeView.hide()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            index = self.indexAt(event.pos())
            self.load_data(index)
        super(MySelectTableView, self).mousePressEvent(event)

    def contextMenuEvent(self, event):
        contextMenu = QMenu(self)

        selected_rows = [index.row() for index in self.selectionModel().selectedRows()]

        if selected_rows:
            action1 = contextMenu.addAction(QCoreApplication.translate('EditorDialog', 'Remove', None),
                                            self.removeAction)
            path = None
            selected_indexes = self.selectionModel().selectedRows()
            if len(selected_indexes) == 1:
                selected_rows = selected_indexes[0].row()
                path = self.model.item(selected_rows, 0).text()
                path = path.replace('\\', '/')
                is_open_filter = self.treeView.model.is_open_filter
            action2 = contextMenu.addAction(QCoreApplication.translate('EditorDialog', 'Export to xlsx file', None),
                                            lambda:
                                            self.export_to_xlsx(path, selected_rows, is_open_filter, self.treeView))
            action3 = contextMenu.addAction(QCoreApplication.translate('EditorDialog', 'Export to html file', None),
                                            lambda:
                                            self.export_to_html(path, selected_rows, is_open_filter, self.treeView))
            action4 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', 'Import html and relative translated contents', None),
                lambda:
                self.import_from_html(path, selected_rows, is_open_filter, self.treeView))
        contextMenu.exec_(event.globalPos())

    def ask_advanced_setting(self, treeView, title, content):
        reply = QMessageBox.question(self,
                                     title,
                                     content,
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        treeView.units_min = 0
        treeView.units_max = sys.maxsize
        treeView.translated_min = 0.0
        treeView.translated_max = 100.0
        editorForm = treeView.tableView.editorForm
        if reply == QMessageBox.Yes:
            if editorForm.myExportXlsxSettingForm is None:
                editorForm.myExportXlsxSettingForm = MyExportSettingForm(parent=editorForm)
            editorForm.myExportXlsxSettingForm.exec()
            if editorForm.myExportXlsxSettingForm.unitsCheckBox.isChecked():
                try:
                    treeView.units_min = int(editorForm.myExportXlsxSettingForm.unitsMinLlineEdit.text())
                except Exception:
                    pass
                try:
                    treeView.units_max = int(editorForm.myExportXlsxSettingForm.unitsMaxLineEdit.text())
                except Exception:
                    pass
            if editorForm.myExportXlsxSettingForm.translatedCheckBox.isChecked():
                try:
                    treeView.translated_min = float(editorForm.myExportXlsxSettingForm.translatedMinLineEdit.text())
                except Exception:
                    pass
                try:
                    treeView.translated_max = float(editorForm.myExportXlsxSettingForm.translatedMaxLineEdit.text())
                except Exception:
                    pass

    def import_from_html(self, path, selected_rows, is_open_filter, treeView: MyTreeView):
        global rpy_info_dic
        if path is None:
            return
        if selected_rows is not None:
            treeView.is_fresh = True
        else:
            treeView.is_fresh = False
        editorForm = treeView.tableView.editorForm
        if editorForm.myImportHtmlForm is None:
            editorForm.myImportHtmlForm = MyImportHtmlForm(parent=self)
        last_write_html = html_util.last_write_html
        if last_write_html is not None:
            editorForm.myImportHtmlForm.selecHtmlFileText.setText(last_write_html)
        editorForm.myImportHtmlForm.exec()
        dic = editorForm.myImportHtmlForm.dic
        treeView.is_replace_special_symbols = editorForm.myImportHtmlForm.is_replace_special_symbols
        if dic is None:
            return
        if os.path.isfile(path):
            if selected_rows is not None:
                self.load_data(self.model.index(selected_rows, 0))
            if path in rpy_info_dic.keys():
                ret, unmatch_cnt, p = rpy_info_dic[path]
            else:
                ret, unmatch_cnt, p = get_rpy_info(path)
                rpy_info_dic[path] = ret, unmatch_cnt, p
            f = io.open(path, 'r', encoding='utf-8')
            _read_lines = f.readlines()
            f.close()
            l = []
            for i, e in enumerate(ret):
                ori_line = e['ori_line'] - 1
                line = e['line'] - 1
                original = e['original']
                current = e['current']
                if treeView.tableView.is_original:
                    target = original
                else:
                    target = current
                replaced = None
                target_key = target
                if treeView.is_replace_special_symbols:
                    d = EncodeBrackets(target)
                    target_key = d['encoded'].strip('"')
                    translated = get_translated(dic, d)
                    replaced = translated
                else:
                    if target_key in dic.keys():
                        replaced = dic[target_key]
                if replaced is None:
                    l.append(target)
                    translated = ''
                    if target_key in dic:
                        translated = dic[target_key]
                    log_print(f'{path} Error in line:{str(line)}\n{target}\n{target_key}\n{translated}\nError')
                    continue
                l.append(replaced)
                if _read_lines[line].startswith('    new '):
                    header = _read_lines[line][:7]
                    content = _read_lines[line][7:]
                    _read_lines[line] = header + content.replace(current, replaced, 1)
                else:
                    _read_lines[line] = _read_lines[line].replace(current, replaced, 1)
            f = io.open(p, 'w', encoding='utf-8')
            f.writelines(_read_lines)
            f.close()
            rpy_info_dic.clear()
            index = QModelIndex(treeView.tableView.index)
            item = treeView.selectTableView.model.item(index.row(), index.column())
            treeView.refresh_table_view(treeView.tableView.file)
            ret, unmatch_cnt, p = rpy_info_dic[treeView.tableView.file]
            if len(ret) != 0:
                percentage = unmatch_cnt / len(ret) * 100
            else:
                percentage = 0
            try:
                if item is not None and item.index() == index:
                    treeView.selectTableView.model.item(index.row(), 2).setText(f'{percentage:.2f}%')
                    treeView.tableView.index = treeView.selectTableView.model.item(index.row(), 2).index()
                else:
                    treeView.model.data(index)
            except Exception as e:
                msg = traceback.format_exc()
                log_print(msg)
        elif os.path.isdir(path):
            treeView.export_path = None
            treeView.export_html = False
            treeView.import_html = True
            self.ask_advanced_setting(treeView, QCoreApplication.translate('EditorDialog', 'Import to files', None),
                                      QCoreApplication.translate('EditorDialog',
                                                                 'Do you want to make advanced settings (the default setting is to import to all files in the directory)',
                                                                 None))
            t = getRpyInfoThread(p=path, is_open_filter=is_open_filter)
            editorForm.setDisabled(True)
            treeView.show()
            t.start()

    def export_to_html(self, path, selected_rows, is_open_filter, treeView):
        global rpy_info_dic
        if path is None:
            return
        if selected_rows is not None:
            treeView.is_fresh = True
        else:
            treeView.is_fresh = False
        if os.path.isfile(path):
            if selected_rows is not None:
                self.load_data(self.model.index(selected_rows, 0))

            fileName, _ = QFileDialog.getSaveFileName(self,
                                                      QCoreApplication.translate('EditorDialog',
                                                                                 'Export to html file',
                                                                                 None), "",
                                                      "Html Files (*.html)")
            if fileName:
                if len(fileName) == 0:
                    return
                if not fileName.endswith('.html'):
                    fileName += '.html'
                if path in rpy_info_dic.keys():
                    ret, unmatch_cnt, p = rpy_info_dic[path]
                else:
                    ret, unmatch_cnt, p = get_rpy_info(path)
                    rpy_info_dic[path] = ret, unmatch_cnt, p
                l = []
                reply = QMessageBox.question(self,
                                             'o((>ω< ))o',
                                             QCoreApplication.translate('EditorDialog',
                                                                        'Do you want to replace special symbols?',
                                                                        None),
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                is_replace_special_symbols = True
                if reply != QMessageBox.Yes:
                    is_replace_special_symbols = False
                for i, e in enumerate(ret):
                    original = e['original']
                    current = e['current']
                    if treeView.tableView.is_original:
                        target = original
                    else:
                        target = current
                    if len(target) > 0:
                        e['target'] = target
                        if is_replace_special_symbols:
                            d = EncodeBrackets(target)
                            if isAllPunctuations(d['encoded'].strip('"')) == False:
                                target = d['encoded'].strip('"')
                                e['target'] = target
                                e['d'] = d
                        ret[i] = e
                        l.append(target)
                data = json.dumps(ret)
                if not is_replace_special_symbols:
                    data = None
                write_html_with_strings(fileName, l, data)
                open_directory_and_select_file(fileName)
        elif os.path.isdir(path):
            fileName, _ = QFileDialog.getSaveFileName(self,
                                                      QCoreApplication.translate('EditorDialog',
                                                                                 'Export to html file',
                                                                                 None), "",
                                                      "Html Files (*.html)")
            if fileName:
                if len(fileName) == 0:
                    return
                if not fileName.endswith('.html'):
                    fileName += '.html'
                treeView.export_path = fileName
                treeView.export_html = True
                treeView.import_html = False
                self.ask_advanced_setting(treeView,
                                          QCoreApplication.translate('EditorDialog', 'Export to html file', None),
                                          QCoreApplication.translate('EditorDialog',
                                                                     'Do you want to make advanced settings (the default setting is to export all files in the directory)',
                                                                     None))
                reply = QMessageBox.question(self,
                                             'o((>ω< ))o',
                                             QCoreApplication.translate('EditorDialog',
                                                                        'Do you want to replace special symbols?',
                                                                        None),
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                treeView.is_replace_special_symbols = True
                if reply != QMessageBox.Yes:
                    treeView.is_replace_special_symbols = False
                t = getRpyInfoThread(p=path, is_open_filter=is_open_filter)
                treeView.tableView.editorForm.setDisabled(True)
                treeView.show()
                t.start()

    def export_to_xlsx(self, path, selected_rows, is_open_filter, treeView):
        global rpy_info_dic
        if path is None:
            return
        if selected_rows is not None:
            treeView.is_fresh = True
        else:
            treeView.is_fresh = False
        if os.path.isfile(path):
            if selected_rows is not None:
                self.load_data(self.model.index(selected_rows, 0))

            fileName, _ = QFileDialog.getSaveFileName(self,
                                                      QCoreApplication.translate('EditorDialog',
                                                                                 'Export to xlsx file',
                                                                                 None), "",
                                                      "Xlsx Files (*.xlsx)")
            if fileName:
                if not fileName.endswith('.xlsx'):
                    fileName += '.xlsx'
                if path in rpy_info_dic.keys():
                    ret, unmatch_cnt, p = rpy_info_dic[path]
                else:
                    ret, unmatch_cnt, p = get_rpy_info(path)
                    rpy_info_dic[path] = ret, unmatch_cnt, p
                wb = Workbook()
                ws = wb.active
                ws.title = "New Sheet"
                ws.cell(row=1, column=1, value=QCoreApplication.translate('EditorDialog', 'Original', None))
                ws.cell(row=1, column=2, value=QCoreApplication.translate('EditorDialog', 'Current', None))

                cnt = 2
                for i in ret:
                    original = i['original']
                    current = i['current']
                    ws.cell(row=cnt, column=1, value=original)
                    ws.cell(row=cnt, column=2, value=current)
                    cnt = cnt + 1
                wb.save(f'{fileName}')
                open_directory_and_select_file(fileName)
        elif os.path.isdir(path):
            fileName, _ = QFileDialog.getSaveFileName(self,
                                                      QCoreApplication.translate('EditorDialog',
                                                                                 'Export to xlsx file',
                                                                                 None), "",
                                                      "Xlsx Files (*.xlsx)")
            if fileName:
                if not fileName.endswith('.xlsx'):
                    fileName += '.xlsx'
                treeView.export_path = fileName
                treeView.export_html = False
                treeView.import_html = False
                self.ask_advanced_setting(treeView,
                                          QCoreApplication.translate('EditorDialog', 'Export to xlsx file', None),
                                          QCoreApplication.translate('EditorDialog',
                                                                     'Do you want to make advanced settings (the default setting is to export all files in the directory)',
                                                                     None))
                t = getRpyInfoThread(p=path, is_open_filter=is_open_filter)
                treeView.tableView.editorForm.setDisabled(True)
                treeView.show()
                t.start()

    def removeAction(self):
        selected_indexes = self.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes]
        selected_rows.sort(reverse=True)
        for row in selected_rows:
            self.tableView.editorForm.addedList.remove(self.model.item(row, 0).text().replace('file:///', ''))
            self.model.removeRow(row)
        self.selectionModel().clearSelection()


class translateThread(threading.Thread):
    def __init__(self, threadID, client, transList, target_language, source_language, fmt, is_open_filter, filter_length, is_replace_special_symbols):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.client = client
        self.transList = transList
        self.target_language = target_language
        self.source_language = source_language
        self.fmt = fmt
        self.is_open_filter = is_open_filter
        self.filter_length = filter_length
        self.is_replace_special_symbols = is_replace_special_symbols


    def run(self):
        global translated_dic, rpy_info_dic, editor_main_trhead_lock, editor_main_thread_tasks
        try:
            log_print('begin translate! please waiting...')
            if isinstance(self.client, str) and self.client == 'web_brower':
                editor_main_trhead_lock.acquire()
                if os.path.isfile(web_brower_export_name):
                    os.remove(web_brower_export_name)
                html_util.plain_text_to_html_from_list(self.transList, web_brower_export_name, self.is_replace_special_symbols)
                import webbrowser
                webbrowser.open(web_brower_export_name)
                dic = dict()
                dic['task'] = 'web_brower_export'
                dic['param'] = self.transList,self.is_open_filter, self.filter_length
                editor_main_thread_tasks.append(dic)
                editor_main_trhead_lock.release()
            else:
                translated_dic = TranslateToList(self.client, self.transList, self.target_language,
                                                 self.source_language,
                                                 fmt=self.fmt)

        except Exception as e:
            msg = traceback.format_exc()
            log_print(msg)
            translated_dic = None


class MyInputDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        layout = QVBoxLayout()
        self.label1 = QLabel()
        self.text_edit = QTextEdit()
        self.text_edit = QTextEdit()
        self.checkbox = QCheckBox()
        self.checkbox.setText(QCoreApplication.translate('EditorDialog', 'Case Sensitive', None))
        self.checkbox.setChecked(True)
        self.referCheckbox = QCheckBox()
        self.referCheckbox.setText(QCoreApplication.translate('EditorDialog', 'Search refer column', None))
        self.referCheckbox.setChecked(True)
        self.originalCheckbox = QCheckBox()
        self.originalCheckbox.setText(QCoreApplication.translate('EditorDialog', 'Search Original column', None))
        self.originalCheckbox.setChecked(True)
        self.currentCheckbox = QCheckBox()
        self.currentCheckbox.setText(QCoreApplication.translate('EditorDialog', 'Search Current column', None))
        self.currentCheckbox.setChecked(True)
        self.translatedCheckbox = QCheckBox()
        self.translatedCheckbox.setText(QCoreApplication.translate('EditorDialog', 'Search Translated column', None))
        self.translatedCheckbox.setChecked(True)
        layout.addWidget(self.label1)
        layout.addWidget(self.text_edit)
        layout.addWidget(self.checkbox)
        layout.addWidget(self.referCheckbox)
        layout.addWidget(self.originalCheckbox)
        layout.addWidget(self.currentCheckbox)
        layout.addWidget(self.translatedCheckbox)
        button = QPushButton('OK')
        button.clicked.connect(self.accept)
        layout.addWidget(button)

        self.setLayout(layout)
        self.setMinimumWidth(1000)

    def getText(self):
        return self.text_edit.toPlainText()


class EnterClickButtonTextEdit(QTextEdit):
    def __init__(self, button: QPushButton):
        super().__init__()
        self.button = button

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Return:
            self.button.click()
        else:
            super().keyPressEvent(event)


class MyInputJumpLineDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout()
        self.label1 = QLabel()
        layout.addWidget(self.label1)
        self.button = QPushButton('OK')
        self.button.clicked.connect(self.accept)
        self.text_edit = EnterClickButtonTextEdit(self.button)
        layout.addWidget(self.text_edit)
        layout.addWidget(self.button)
        self.setLayout(layout)

    def getText(self):
        return self.text_edit.toPlainText().strip()


class MyTableView(QTableView):
    def __init__(self, parent=None):
        super(MyTableView, self).__init__(parent)
        self.model = QStandardItemModel()
        self.setModel(self.model)
        self.model.setHorizontalHeaderLabels([QCoreApplication.translate('EditorDialog', 'line', None),
                                              QCoreApplication.translate('EditorDialog', 'refer', None),
                                              QCoreApplication.translate('EditorDialog', 'Original', None),
                                              QCoreApplication.translate('EditorDialog', 'Current', None),
                                              QCoreApplication.translate('EditorDialog', 'Translated', None)])
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.verticalHeader().setVisible(False)
        self.setSelectionBehavior(QTableView.SelectRows)
        self.is_original = False
        self.model.dataChanged.connect(self.handle_data_changed)
        self.file = None
        self.row = None
        self.searched = set()
        self.local_glossary = None
        selection_model = self.selectionModel()
        selection_model.selectionChanged.connect(self.on_selection_changed)

    def on_selection_changed(self, selected, deselected):
        indexes = self.selectionModel().selectedRows()
        indexes = sorted(indexes, key=lambda index: index.row())
        l = []
        if indexes:
            for i in indexes:
                index = QModelIndex(i)
                if (self.copy_index + 1) > 0:
                    data = self.model.index(index.row(), self.copy_index + 1).data()
                    l.append(data)
        if len(l) > 0:
            text_to_copy = "\n".join(l)
            pyperclip.copy(text_to_copy)

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Find):
            self.search()
        elif event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_G:
            self.jump_line()
        else:
            super().keyPressEvent(event)

    def jump_line(self):
        dialog = MyInputJumpLineDialog(self)
        dialog.setWindowTitle(QCoreApplication.translate('EditorDialog', 'Input Dialog', None))
        dialog.label1.setText(
            QCoreApplication.translate('EditorDialog', 'Please Input the line number you want to jump', None))
        if dialog.exec():
            text = dialog.getText()
            for row in range(self.model.rowCount()):
                item = self.model.item(row, 0)
                if item and item.text() == text:
                    self.selectRow(row)

    def search(self):
        dialog = MyInputDialog(self)
        dialog.setWindowTitle(QCoreApplication.translate('EditorDialog', 'Input Dialog', None))
        dialog.label1.setText(
            QCoreApplication.translate('EditorDialog', 'Please Input the content you want to search', None))
        if dialog.exec():
            self.searched.clear()
            search_text = dialog.getText()
            if len(search_text) == 0:
                return
            match_count = 0
            for row in range(self.model.rowCount()):
                is_match = False
                for column in range(self.model.columnCount()):
                    if column == 0:
                        continue
                    if not dialog.referCheckbox.isChecked() and column == 1:
                        continue
                    if not dialog.originalCheckbox.isChecked() and column == 2:
                        continue
                    if not dialog.currentCheckbox.isChecked() and column == 3:
                        continue
                    if not dialog.translatedCheckbox.isChecked() and column == 4:
                        continue
                    if self.verticalHeader().isSectionHidden(row):
                        continue
                    item = self.model.item(row, column)
                    if item:
                        item_text = item.text()
                        if not dialog.checkbox.isChecked():
                            item_text = item_text.lower()
                            search_text = search_text.lower()
                        if search_text in item_text:
                            self.selectRow(row)
                            self.searched.add(row)
                            is_match = True
                            log_print(f'search match line:{row} column{column} content:{item.text()}')
                            match_count = match_count + 1
                if not is_match:
                    self.verticalHeader().setSectionHidden(row, True)
                else:
                    self.verticalHeader().setSectionHidden(row, False)
            if match_count == 0:
                log_print('search not found!')
            else:
                log_print(f'search match count:{match_count}')
            self.editorForm.searchedOnlyCheckBox.setChecked(True)
            if self.editorForm.logAfterSearchCheckBox.isChecked():
                self.editorForm.parent.showNormal()
                self.editorForm.parent.raise_()

    def handle_data_changed(self, top_left, bottom_right):
        top_row = top_left.row()
        bottom_row = bottom_right.row()
        left_column = top_left.column()
        right_column = bottom_right.column()

        for row in range(top_row, bottom_row + 1):
            for column in range(left_column, right_column + 1):
                index = self.model.index(row, 0)
                now_data = self.model.index(row, column).data()
                item = QStandardItem(index.model().itemFromIndex(index))
                new_data = item.data(Qt.UserRole)
                tmp_data = dict()
                tmp_data['original'] = new_data['original']
                tmp_data['current'] = new_data['current']
                if column == 0:
                    tmp_data['line'] = int(now_data)
                elif column == 1:
                    tmp_data['refer'] = now_data
                elif column == 2:
                    tmp_data['original'] = now_data
                elif column == 3:
                    tmp_data['current'] = now_data
                elif column == 4:
                    tmp_data['translated'] = now_data
                new_data['is_match'] = tmp_data['original'] == tmp_data['current']
                self.model.dataChanged.disconnect()
                item.setToolTip(now_data)
                column_count = self.model.columnCount()
                item.setData(new_data, Qt.UserRole)
                self.model.setItem(row, 0, item)
                self.model.dataChanged.connect(self.handle_data_changed)

    def contextMenuEvent(self, event):
        contextMenu = QMenu(self)

        selected_rows = [index.row() for index in self.selectionModel().selectedRows()]

        if selected_rows:
            action1 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', "Translate Translation Source to Translated", None),
                self.translate)
            action2 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', "Copy Original to Current", None), self.copy_ori_to_cur)
            action3 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', "Copy Translated to Current", None),
                self.copy_translated_to_cur)
            action4 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', "Rollback Current to First Load", None), self.rollback_cur)
            action5 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', "Export to xlsx file", None), self.export_to_xlsx)
            action6 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', "Export to html file", None), self.export_to_html)
            action7 = contextMenu.addAction(
                QCoreApplication.translate('EditorDialog', 'Import html and relative translated contents', None),
                self.import_from_html)

        contextMenu.exec_(event.globalPos())

    def import_from_html(self):
        selected_indexes = self.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
        selected_rows.sort(reverse=False)
        editorForm = self.editorForm
        if editorForm.myImportHtmlForm is None:
            editorForm.myImportHtmlForm = MyImportHtmlForm(parent=self)
        last_write_html = html_util.last_write_html
        if last_write_html is not None:
            editorForm.myImportHtmlForm.selecHtmlFileText.setText(last_write_html)
        editorForm.myImportHtmlForm.exec()
        dic = editorForm.myImportHtmlForm.dic
        is_replace_special_symbols = editorForm.myImportHtmlForm.is_replace_special_symbols
        if dic is None:
            return
        l = []
        for row in selected_rows:
            original = self.model.item(row, 0).data(Qt.UserRole)['original']
            # current = self.model.item(row, 0).data(Qt.UserRole)['current']
            current = self.model.item(row, 3).text()
            ori_line = self.model.item(row, 0).data(Qt.UserRole)['ori_line'] - 1
            line = self.model.item(row, 0).data(Qt.UserRole)['line'] - 1
            if self.is_original:
                target = original
            else:
                target = current
            replaced = None
            target_key = target
            if is_replace_special_symbols:
                d = EncodeBrackets(target)
                target_key = d['encoded'].strip('"')
                translated = get_translated(dic, d)
                replaced = translated
            else:
                if target_key in dic.keys():
                    replaced = dic[target_key]
            if replaced is None:
                l.append(target)
                continue
            l.append(replaced)
            self.model.item(row, 4).setText(replaced)
        if editorForm.autoCopyCheckBox.isChecked():
            self.copy_translated_to_cur()

    def export_to_html(self):
        fileName, _ = QFileDialog.getSaveFileName(self,
                                                  QCoreApplication.translate('EditorDialog', "Export to html file",
                                                                             None), "",
                                                  "Html Files (*.html)")
        if len(fileName) == 0:
            return
        if not fileName.endswith('.html'):
            fileName += '.html'
        reply = QMessageBox.question(self,
                                     'o((>ω< ))o',
                                     QCoreApplication.translate('EditorDialog',
                                                                'Do you want to replace special symbols?',
                                                                None),
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        is_replace_special_symbols = True
        if reply != QMessageBox.Yes:
            is_replace_special_symbols = False
        selected_indexes = self.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
        selected_rows.sort(reverse=False)
        l = []
        ret_l = []
        cnt = 0
        for row in selected_rows:
            e = self.model.item(row, 0).data(Qt.UserRole)
            original = self.model.item(row, 0).data(Qt.UserRole)['original']
            # current = self.model.item(row, 0).data(Qt.UserRole)['current']
            current = self.model.item(row, 3).text()
            if self.is_original:
                target = original
            else:
                target = current
            if len(target) > 0:
                e['target'] = target
                if is_replace_special_symbols:
                    d = EncodeBrackets(target)
                    if isAllPunctuations(d['encoded'].strip('"')) == False:
                        target = d['encoded'].strip('"')
                        e['target'] = target
                        e['d'] = d
                l.append(target)
                ret_l.append(e)
                cnt = cnt + 1
        data = json.dumps(ret_l)
        if not is_replace_special_symbols:
            data = None
        write_html_with_strings(fileName, l, data)
        open_directory_and_select_file(fileName)

    def export_to_xlsx(self):
        fileName, _ = QFileDialog.getSaveFileName(self,
                                                  QCoreApplication.translate('EditorDialog', "Export to xlsx file",
                                                                             None), "",
                                                  "Xlsx Files (*.xlsx)")
        if len(fileName) == 0:
            return
        if fileName:
            if not fileName.endswith('.xlsx'):
                fileName += '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = "New Sheet"
            ws.cell(row=1, column=1, value=self.model.horizontalHeaderItem(2).text())
            ws.cell(row=1, column=2, value=self.model.horizontalHeaderItem(3).text())
            selected_indexes = self.selectionModel().selectedRows()
            selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
            selected_rows.sort(reverse=False)
            cnt = 2
            for row in selected_rows:
                original = self.model.item(row, 0).data(Qt.UserRole)['original']
                current = self.model.item(row, 0).data(Qt.UserRole)['current']
                ws.cell(row=cnt, column=1, value=original)
                ws.cell(row=cnt, column=2, value=current)
                cnt = cnt + 1
            wb.save(f'{fileName}')
            open_directory_and_select_file(fileName)

    def rollback_cur(self):
        selected_indexes = self.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
        selected_rows.sort(reverse=True)
        for row in selected_rows:
            current = self.model.item(row, 0).data(Qt.UserRole)['current']
            self.model.item(row, 3).setText(current)

    def translate(self):
        if os.path.isfile(web_brower_export_name):
            os.remove(web_brower_export_name)
        selected_indexes = self.selectionModel().selectedRows()
        self.selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
        self.selected_rows.sort(reverse=True)
        local_glossary = self.local_glossary
        transList = []
        client = init_client()
        is_replace_special_symbols = self.editorForm.replaceCheckBox.isChecked()
        if isinstance(client, str) and client == 'web_brower':
            target_language = None
            source_language = None
        else:
            target_language = targetDic[self.editorForm.targetComboBox.currentText()]
            source_language = sourceDic[self.editorForm.sourceComboBox.currentText()]
        self.editorForm.treeView.is_replace_special_symbols = is_replace_special_symbols
        for row in self.selected_rows:
            if self.is_original:
                target = self.model.item(row, 2).text()
            else:
                target = self.model.item(row, 3).text()
            if len(target) == 0:
                line_index = int(self.model.item(row, 0).text())
                log_print(
                    'Empty in line:' + str(line_index) + ' ' + ' Skip Translation\n')
                continue
            if local_glossary is not None:
                for original, replace in local_glossary.items():
                    target = target.replace(original, replace)
            if is_replace_special_symbols:
                d = EncodeBrackets(target)
            else:
                d = dict()
                d['en_1'] = []
                d['en_2'] = []
                d['en_3'] = []
                d['encoded'] = target
            strip_i = target
            for j in (d['en_1']):
                strip_i = strip_i.replace(j, '')
            for j in (d['en_2']):
                strip_i = strip_i.replace(j, '')
            for j in (d['en_3']):
                strip_i = strip_i.replace(j, '')
            _strip_i = replace_all_blank(strip_i)
            self.editorForm : MyEditorForm
            filter_length = self.editorForm.filterLengthLineEdit.text()
            if len(filter_length) == 0:
                filter_length = '0'
            filter_length = int(filter_length)
            if self.editorForm.filterCheckBox.isChecked():
                if len(_strip_i) < filter_length:
                    # log_print(len(strip_i),i)
                    continue
            if (isAllPunctuations(d['encoded'].strip('"')) == False):
                if is_replace_special_symbols:
                    transList.append(d['encoded'].strip('"'))
                else:
                    transList.append(target)
        if len(transList) == 0:
            log_print('The translation content is empty!')
            self.editorForm.parent.showNormal()
            self.editorForm.parent.raise_()
            return

        if client is None:
            return
        if client.__class__.__name__ == 'Translate' and local_glossary is not None and len(local_glossary) > 0:
            fmt = 'html'
        else:
            fmt = 'text'
        self.editorForm.parent.showNormal()
        self.editorForm.parent.raise_()
        self.editorForm.hide()
        self.editorForm.setDisabled(True)
        # trans_dic = TranslateToList(client, transList, target_language, source_language)
        global translated_thread, translated_dic
        translated_thread = translateThread(0, client, transList, target_language, source_language, fmt=fmt, is_open_filter=self.editorForm.filterCheckBox.isChecked(), filter_length= filter_length, is_replace_special_symbols = is_replace_special_symbols)
        translated_thread.start()

    def copy_translated_to_cur(self):
        selected_indexes = self.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
        selected_rows.sort(reverse=True)
        for row in selected_rows:
            translated = self.model.item(row, 4).text()
            self.model.item(row, 3).setText(translated)

    def copy_ori_to_cur(self):
        selected_indexes = self.selectionModel().selectedRows()
        selected_rows = [index.row() for index in selected_indexes if not self.editorForm.tableView.verticalHeader().isSectionHidden(index.row())]
        selected_rows.sort(reverse=True)
        for row in selected_rows:
            ori = self.model.item(row, 2).text()
            self.model.item(row, 3).setText(ori)


class MyEditorForm(QDialog, Ui_EditorDialog):
    def __init__(self, parent=None):
        super(MyEditorForm, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('main.ico'))
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinMaxButtonsHint)
        self.addedList = []
        self.selectFilesBtn.clicked.connect(self.select_file)
        self.selectDirBtn.clicked.connect(self.select_directory)
        self.addListButton.clicked.connect(self.add_to_list)
        self.tableView = MyTableView()
        self.tableView.editorForm = self
        self.tableVerticalLayout.addWidget(self.tableView)
        self.selectTableView = MySelectTableView()
        self.fileVerticalLayout.addWidget(self.selectTableView)
        self.treeView = MyTreeView()
        self.fileVerticalLayout.addWidget(self.treeView)
        self.treeView.setMaximumWidth(550)
        self.selectTableView.setMaximumWidth(550)
        self.selectTableView.tableView = self.tableView
        self.selectTableView.treeView = self.treeView
        self.treeView.tableView = self.tableView
        self.treeView.selectTableView = self.selectTableView
        self.rpyCheckBox.setChecked(True)
        self.selectTableView.rpyCheckBox = self.rpyCheckBox
        self.rpyCheckBox.stateChanged.connect(self.on_rpy_checkbox_state_changed)
        self.treeView.hide()
        self.buttonGroup = QButtonGroup()
        self.buttonGroup.addButton(self.originalRadioButton, 1)
        self.buttonGroup.addButton(self.currentRadioButton, 2)
        self.buttonGroup.buttonClicked.connect(self.button_group_clicked)
        self.buttonGroup2 = QButtonGroup()
        self.buttonGroup2.addButton(self.copyOriginalRadioButton, 1)
        self.buttonGroup2.addButton(self.copyCurrentRadioButton, 2)
        self.buttonGroup2.addButton(self.copyTranslatedRadioButton, 3)
        self.buttonGroup2.buttonClicked.connect(self.button_group_clicked2)
        self.copyCurrentRadioButton.setChecked(True)
        self.tableView.copy_index = -2
        self.copySelectedCheckBox.stateChanged.connect(self.on_copy_selected_checkbox_state_changed)
        self.init_combobox()
        self.untranslatedCheckBox.stateChanged.connect(self.on_checkbox_state_changed)
        self.saveFileButton.clicked.connect(self.save_file_button_clicked)
        self.changeTranslationEngineButton.clicked.connect(self.show_engine_settings)
        self.showLogButton.clicked.connect(self.on_show_log_button_checked)
        self.searchedOnlyCheckBox.stateChanged.connect(self.on_checkbox_state_changed)
        self.localGlossaryCheckBox.clicked.connect(self.on_local_glossary_checkbox_state_changed)
        self.myImportHtmlForm = None
        self.myExportXlsxSettingForm = None
        self.filterLengthLineEdit.setValidator(QIntValidator(1, 99, self))
        _thread.start_new_thread(self.update_log, ())

    def on_local_glossary_checkbox_state_changed(self):
        self.tableView.local_glossary = None
        if self.localGlossaryCheckBox.isChecked():
            local_glossary_form = self.parent.local_glossary_form
            local_glossary_form.exec()
            dic = local_glossary_form.data
            index = self.sourceComboBox.findText('Auto Detect')
            if dic is None or len(dic) == 0:
                self.localGlossaryCheckBox.setChecked(False)
                if 'Auto Detect' in sourceDic.keys() and index == -1:
                    self.sourceComboBox.addItem('Auto Detect')
                    index = self.sourceComboBox.findText('Auto Detect')
                    self.sourceComboBox.setCurrentIndex(index)
            else:
                if index != -1:
                    current_index = self.sourceComboBox.currentIndex()
                    self.sourceComboBox.removeItem(index)
                    if current_index == index:
                        self.sourceComboBox.setCurrentIndex(0)
                self.tableView.local_glossary = dic
        else:
            index = self.sourceComboBox.findText('Auto Detect')
            if 'Auto Detect' in sourceDic.keys() and index == -1:
                self.sourceComboBox.addItem('Auto Detect')
                index = self.sourceComboBox.findText('Auto Detect')
                self.sourceComboBox.setCurrentIndex(index)

    def on_copy_selected_checkbox_state_changed(self):
        if self.copySelectedCheckBox.isChecked():
            self.tableView.copy_index = self.buttonGroup2.checkedId()
        else:
            self.tableView.copy_index = -2

    def on_show_log_button_checked(self):
        self.parent.showNormal()
        self.parent.raise_()

    def closeEvent(self, event):
        global rpy_info_dic
        rpy_info_dic.clear()
        self.parent.widget.show()
        self.parent.menubar.show()
        self.parent.versionLabel.show()
        self.parent.actionedit.triggered.connect(lambda: self.parent.show_edit_form())
        self.parent.init_combobox()
        self.parent.showNormal()
        self.hide()
        event.ignore()
        return

    def show_engine_settings(self):
        engine_form = MyEngineForm(parent=self)
        ori = None
        now = None
        if os.path.isfile('engine.txt'):
            with open('engine.txt', 'r') as json_file:
                ori = json.load(json_file)
        engine_form.exec()
        if os.path.isfile('engine.txt'):
            with open('engine.txt', 'r') as json_file:
                now = json.load(json_file)
        if (now is not None and ori is not None and now['engine'] != ori['engine']) \
                or (now is None and ori is None) \
                or (now is not None and ori is None):
            self.init_combobox()

    def save_file_button_clicked(self):
        if self.tableView.file is None:
            return
        log_print('start saving file! please waiting...')
        self.hide()
        self.setDisabled(True)
        self.parent.showNormal()
        self.parent.raise_()
        global is_need_save
        is_need_save = True

    def on_checkbox_state_changed(self):
        model = self.tableView.model
        rows = model.rowCount()
        columns = model.columnCount()
        for row in range(rows):
            for column in range(columns):
                index = model.index(row, 0)
                item = index.model().itemFromIndex(index)
                data = item.data(Qt.UserRole)
                original = data['original']
                current = self.tableView.model.item(row, 3).text()
                if current == '' and self.untranslatedCheckBox.isChecked():
                    self.tableView.verticalHeader().setSectionHidden(row, False)
                elif original!=current and self.untranslatedCheckBox.isChecked():
                    self.tableView.verticalHeader().setSectionHidden(row, True)
                else:
                    self.tableView.verticalHeader().setSectionHidden(row, False)

                if self.tableView.verticalHeader().isSectionHidden(row):
                    continue

                if len(self.tableView.searched) > 0 and self.searchedOnlyCheckBox.isChecked():
                    if row not in self.tableView.searched:
                        self.tableView.verticalHeader().setSectionHidden(row, True)
                    else:
                        self.tableView.verticalHeader().setSectionHidden(row, False)
                # log_print(f"Row: {row}, Column: {column}, Data: {data}")

    @staticmethod
    def get_combobox_content(p, d):
        if not os.path.isfile(p):
            return []
        try:
            f = io.open(p, 'r', encoding='utf-8')
            _read = f.read()
            f.close()
            if len(_read) == 0:
                return []
            _read_line = _read.split('\n')
            ret_l = []
            for i in _read_line:
                contents = i.split(':')
                d[contents[0].strip()] = contents[1].strip()
                ret_l.append(contents[0].strip())
            ret_l = list(set(ret_l))
            ret_l.sort()
            return ret_l
        except:
            msg = traceback.format_exc()
            log_print(msg)
            return []

    def on_combobox_changed(self):
        if os.path.isfile('engine.txt'):
            json_file = open('engine.txt', 'r',encoding='utf-8')
            ori = json.load(json_file)
            json_file.close()
            current_engine = ori['engine']
            dic = dict()
            dic['target'] = self.targetComboBox.currentText()
            dic['source'] = self.sourceComboBox.currentText()
            ori[current_engine] = dic
            json_file = open('engine.txt', 'w', encoding='utf-8')
            json.dump(ori, json_file)

    def init_combobox(self):
        self.targetComboBox.currentTextChanged.connect(self.on_combobox_changed)
        self.sourceComboBox.currentTextChanged.connect(self.on_combobox_changed)
        self.targetComboBox.currentTextChanged.disconnect()
        self.sourceComboBox.currentTextChanged.disconnect()
        self.targetComboBox.clear()
        self.sourceComboBox.clear()
        targetDic.clear()
        sourceDic.clear()
        target = 'google.target.rst'
        source = 'google.source.rst'
        customEngineDic = dict()
        if os.path.isfile('custom.txt'):
            f = io.open('custom.txt', 'r', encoding='utf-8')
            customEngineDic = json.load(f)
            f.close()
        if os.path.isfile('engine.txt'):
            with open('engine.txt', 'r') as json_file:
                loaded_data = json.load(json_file)
                if loaded_data['engine'] in engineDic:
                    target = engineDic[loaded_data['engine']]['target']
                    source = engineDic[loaded_data['engine']]['source']
                elif loaded_data['engine'] in customEngineDic:
                    target = customEngineDic[loaded_data['engine']]['target']
                    source = customEngineDic[loaded_data['engine']]['source']
                else:
                    log_print(loaded_data['engine'] + ' not in dic')
            if target is None or source is None:
                log_print('target or source not found!')
                return
            if len(target) == 0 or len(source) == 0:
                log_print('target or source is empty!')
                return
        target = language_header + target
        source = language_header + source
        target_l = self.get_combobox_content(target, targetDic)
        for i in target_l:
            self.targetComboBox.addItem(i)
        source_l = self.get_combobox_content(source, sourceDic)
        for i in source_l:
            self.sourceComboBox.addItem(i)
        if self.localGlossaryCheckBox.isChecked():
            index = self.sourceComboBox.findText('Auto Detect')
            if index != -1:
                self.sourceComboBox.removeItem(index)
        try:
            self.sourceComboBox.setCurrentIndex(source_l.index('Auto Detect'))
        except Exception:
            pass
        if os.path.isfile('engine.txt'):
            json_file = open('engine.txt', 'r', encoding='utf-8')
            json_data = json.load(json_file)
            json_file.close()
            current_engine = json_data['engine']
            if current_engine in json_data:
                combobox_data = json_data[current_engine]
                if 'source' in combobox_data:
                    try:
                        self.sourceComboBox.setCurrentIndex(source_l.index(combobox_data['source']))
                    except:
                        pass
                if 'target' in combobox_data:
                    try:
                        self.targetComboBox.setCurrentIndex(target_l.index(combobox_data['target']))
                    except:
                        pass
        self.targetComboBox.currentTextChanged.connect(self.on_combobox_changed)
        self.sourceComboBox.currentTextChanged.connect(self.on_combobox_changed)

    def button_group_clicked2(self, item):
        self.tableView.copy_index = item.group().checkedId()

    def button_group_clicked(self, item):
        if item.group().checkedId() == 1:
            self.tableView.is_original = True
        else:
            self.tableView.is_original = False

    def on_rpy_checkbox_state_changed(self):
        if self.rpyCheckBox.isChecked():
            self.treeView.model.is_open_filter = True
        else:
            self.treeView.model.is_open_filter = False
        indexes = self.selectTableView.selectionModel().selectedIndexes()
        if indexes is not None and len(indexes) > 0:
            self.selectTableView.load_data(indexes[0])

    def add_to_list(self):
        select_files = self.selectFilesText.toPlainText().split('\n')
        cnt = 0
        for i in select_files:
            i = i.replace('file:///', '')
            if len(i) > 0 and os.path.isfile(i):
                if i not in self.addedList:
                    ret, unmatch_cnt, p = get_rpy_info(i)
                    if len(ret) != 0:
                        percentage = unmatch_cnt / len(ret) * 100
                    else:
                        percentage = 0
                    row = [
                        QStandardItem(i),
                        QStandardItem(str(len(ret))),
                        QStandardItem(f'{percentage:.2f}%'),
                    ]
                    for item in row:
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setEditable(False)
                        item.setToolTip(item.text())
                    row[0].setTextAlignment(Qt.AlignLeft)
                    self.selectTableView.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
                    self.selectTableView.model.appendRow(row)
                    self.addedList.append(i)
        self.selectTableView.model.sort(0, Qt.AscendingOrder)

        select_dir = self.selectDirText.toPlainText()
        select_dir = select_dir.replace('file:///', '')
        if len(select_dir) > 0 and os.path.isdir(select_dir):
            if select_dir not in self.addedList:
                row = [
                    QStandardItem(select_dir),
                    QStandardItem('/'),
                    QStandardItem('/'),
                ]
                for item in row:
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setEditable(False)
                    item.setToolTip(item.text())
                row[0].setTextAlignment(Qt.AlignLeft)
                self.selectTableView.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
                self.selectTableView.model.appendRow(row)
                self.addedList.append(select_dir)
                self.selectTableView.model.sort(0, Qt.AscendingOrder)

    def select_directory(self):
        directory = QFileDialog.getExistingDirectory(self, QCoreApplication.translate('EditorDialog',
                                                                                      'select the directory you want to edit',
                                                                                      None))
        self.selectDirText.setText(directory)

    def select_file(self):
        files, filetype = QFileDialog.getOpenFileNames(self,
                                                       QCoreApplication.translate('EditorDialog',
                                                                                  'select the file(s) you want to edit',
                                                                                  None),
                                                       '',
                                                       "Rpy Files (*.rpy);;All Files (*)")
        s = ''
        for file in files:
            s = s + file + '\n'
        self.selectFilesText.setText(s.rstrip('\n'))

    def update_log(self):
        thread = self.UpdateThread()
        thread.update_date.connect(self.update_progress)
        while True:
            thread.start()
            time.sleep(0.5)

    def update_progress(self):
        try:
            global editor_main_trhead_lock, editor_main_thread_tasks, rpy_info_dic
            global translated_thread, translated_dic
            editor_main_trhead_lock.acquire()
            if len(editor_main_thread_tasks) > 0:
                task_info = editor_main_thread_tasks[0]
                editor_main_trhead_lock.release()
                task = task_info['task']
                editor_main_thread_tasks.remove(task_info)
                if task == 'web_brower_export':
                    translate_list, is_open_filter, filter_length = task_info['param']
                    translated_form = MyTranslatedForm()
                    translated_form.exec()
                    f = io.open('translated.txt', 'w', encoding='utf-8')
                    f.write(translated_form.plainTextEdit.toPlainText())
                    f.close()
                    translated_dic, is_replace_special_symbols = get_translated_dic(web_brower_export_name, 'translated.txt')
                    self.treeView.is_replace_special_symbols = is_replace_special_symbols
                    if translated_dic is None:
                        msg_box = QMessageBox()
                        msg_box.setWindowTitle('o(≧口≦)o')
                        msg_box.setText(
                            QCoreApplication.translate('ImportHtmlDialog',
                                                       'The html file does not match the translated file , please check the input files',
                                                       None))
                        msg_box.exec()
                        rpy_info_dic.clear()
                        self.is_waiting_translated = False
                        translated_dic = []
                    rpy_info_dic.clear()
                    if os.path.isfile(web_brower_export_name):
                        os.remove(web_brower_export_name)
            else:
                editor_main_trhead_lock.release()

            rpy_lock.acquire()
            if os.path.isfile('rpy_info_got') and os.path.getsize('rpy_info_got') > 0:
                f = io.open('rpy_info_got', 'r', encoding='utf-8')
                select_one = f.read()
                f.close()
                if os.path.isdir(select_one):
                    is_fresh = True
                    if self.treeView.import_html:
                        last_translated_dic = html_util.last_translated_dic
                        last_write_html = html_util.last_write_html
                        if last_translated_dic is not None:
                            is_fresh = self.treeView.is_fresh
                            paths = os.walk(select_one, topdown=False)
                            select_one = select_one.replace('\\', '/')
                            l = []
                            for path, dir_lst, file_lst in paths:
                                for file_name in file_lst:
                                    if self.treeView.model.is_open_filter and not file_name.endswith("rpy"):
                                        continue
                                    i = os.path.join(path, file_name)
                                    i = i.replace('\\', '/')
                                    # log_print(i)
                                    ret, unmatch_cnt, p = rpy_info_dic[i]
                                    units = len(ret)
                                    if units != 0:
                                        percentage = unmatch_cnt / units * 100
                                    else:
                                        percentage = 0
                                    if percentage < self.treeView.translated_min or percentage > self.treeView.translated_max:
                                        continue
                                    if units < self.treeView.units_min or units > self.treeView.units_max:
                                        continue
                                    f = io.open(i, 'r', encoding='utf-8')
                                    _read_lines = f.readlines()
                                    f.close()

                                    for index, dic in enumerate(ret):
                                        ori_line = dic['ori_line'] - 1
                                        line = dic['line'] - 1
                                        original = dic['original']
                                        current = dic['current']
                                        is_match = dic['is_match']
                                        if self.treeView.tableView.is_original:
                                            target = original
                                        else:
                                            target = current
                                        replaced = None
                                        target_key = target
                                        if self.treeView.is_replace_special_symbols:
                                            d = EncodeBrackets(target)
                                            target_key = d['encoded'].strip('"')
                                            translated = get_translated(dic, d)
                                            replaced = translated
                                        else:
                                            if target_key in last_translated_dic.keys():
                                                replaced = last_translated_dic[target_key]
                                        if replaced is None:
                                            translated = ''
                                            if target_key in dic:
                                                translated = dic[target_key]
                                            log_print(
                                                f'{i} Error in line:{str(line)}\n{target}\n{target_key}\n{translated}\nError')
                                            continue
                                        dic['current'] = replaced
                                        is_match_now = replaced == original
                                        if is_match_now != is_match:
                                            if is_match_now:
                                                unmatch_cnt = unmatch_cnt - 1
                                            else:
                                                unmatch_cnt = unmatch_cnt + 1
                                        dic['is_match'] = is_match_now
                                        ret[index] = dic
                                        rpy_info_dic[i] = ret, unmatch_cnt, p
                                        if _read_lines[line].startswith('    new '):
                                            header = _read_lines[line][:7]
                                            content = _read_lines[line][7:]
                                            _read_lines[line] = header + content.replace(current, replaced, 1)
                                        else:
                                            _read_lines[line] = _read_lines[line].replace(current, replaced, 1)
                                    f = io.open(i, 'w', encoding='utf-8')
                                    f.writelines(_read_lines)
                                    f.close()
                                    ret, unmatch_cnt, p = rpy_info_dic[i]
                                    for i in ret:
                                        original = i['original']
                                        current = i['current']
                                        if self.tableView.is_original:
                                            target = original
                                        else:
                                            target = current
                                        if len(target) > 0:
                                            l.append(target)
                        self.treeView.import_html = False
                    if self.treeView.export_path is not None:
                        is_fresh = self.treeView.is_fresh
                        paths = os.walk(select_one, topdown=False)
                        select_one = select_one.replace('\\', '/')
                        if self.treeView.export_html == False:
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "New Sheet"
                            ws.cell(row=1, column=1, value=QCoreApplication.translate('EditorDialog', 'Original', None))
                            ws.cell(row=1, column=2, value=QCoreApplication.translate('EditorDialog', 'Current', None))
                            cnt = 2
                        fileName = self.treeView.export_path
                        l = []
                        ret_l = []
                        for path, dir_lst, file_lst in paths:
                            for file_name in file_lst:
                                if self.treeView.model.is_open_filter and not file_name.endswith("rpy"):
                                    continue
                                i = os.path.join(path, file_name)
                                i = i.replace('\\', '/')
                                # log_print(i)
                                ret, unmatch_cnt, p = rpy_info_dic[i]
                                units = len(ret)
                                if units != 0:
                                    percentage = unmatch_cnt / units * 100
                                else:
                                    percentage = 0
                                if percentage < self.treeView.translated_min or percentage > self.treeView.translated_max:
                                    continue
                                if units < self.treeView.units_min or units > self.treeView.units_max:
                                    continue
                                for i, e in enumerate(ret):
                                    original = e['original']
                                    current = e['current']
                                    if self.treeView.export_html == False:
                                        ws.cell(row=cnt, column=1, value=original)
                                        ws.cell(row=cnt, column=2, value=current)
                                        cnt = cnt + 1
                                    else:
                                        if self.tableView.is_original:
                                            target = original
                                        else:
                                            target = current
                                        if len(target) > 0:
                                            e['target'] = target
                                            if self.treeView.is_replace_special_symbols:
                                                d = EncodeBrackets(target)
                                                if isAllPunctuations(d['encoded'].strip('"')) == False:
                                                    target = d['encoded'].strip('"')
                                                    e['target'] = target
                                                    e['d'] = d
                                            ret[i] = e
                                            l.append(target)
                                ret_l.extend(ret)
                        if self.treeView.export_html == False:
                            wb.save(f'{fileName}')
                        else:
                            data = json.dumps(ret_l)
                            if not self.treeView.is_replace_special_symbols:
                                data = None
                            write_html_with_strings(fileName, l, data)
                        open_directory_and_select_file(fileName)

                        self.treeView.export_path = None
                        self.treeView.export_html = False

                    if is_fresh:
                        self.treeView.proxy_model.invalidateFilter()
                        self.treeView.model.setRootPath(select_one)
                        self.treeView.setRootIndex(
                            self.treeView.proxy_model.mapFromSource(self.treeView.model.index(select_one)))
                    if self.tableView.file is not None:
                        index = QModelIndex(self.tableView.index)
                        item = self.selectTableView.model.item(index.row(), index.column())
                        self.treeView.refresh_table_view(self.tableView.file)
                    self.setEnabled(True)
                    os.remove('rpy_info_got')
            rpy_lock.release()
            global is_need_save
            if is_need_save:
                is_need_save = False
                f = io.open(self.tableView.file, 'r', encoding='utf-8')
                _read_lines = f.readlines()
                f.close()
                for row in range(self.tableView.model.rowCount()):
                    data = self.tableView.model.item(row, 0).data(Qt.UserRole)
                    line = int(data['line']) - 1
                    ori_line = int(data['ori_line']) - 1
                    ori_current = str(data['current'])

                    current = self.tableView.model.item(row, 3).text()
                    if ori_current != current:
                        if ori_current != '':
                            if _read_lines[line].startswith('    new '):
                                header = _read_lines[line][:7]
                                content = _read_lines[line][7:]
                                _read_lines[line] = header + content.replace(ori_current, current, 1)
                            else:
                                _read_lines[line] = _read_lines[line].replace(ori_current, current, 1)
                        else:
                            if _read_lines[ori_line].startswith('    old '):
                                _read_lines[line] = '    new ' + '"' + current + '"' + '\n'
                            else:
                                is_empty = bool(data['is_empty'])
                                if is_empty:
                                    ori_content = str(data['ori_content'])
                                    current_content = str(data['current_content'])
                                    _read_lines[line] = current_content.rstrip('""') + f'"{current}"' + '\n'
                                else:
                                    _read_lines[line] = '    ' + '"' + current + '"' + '\n'
                f = io.open(self.tableView.file, 'w', encoding='utf-8')
                f.writelines(_read_lines)
                f.close()
                rpy_info_dic.clear()
                index = QModelIndex(self.tableView.index)
                item = self.selectTableView.model.item(index.row(), index.column())
                self.treeView.refresh_table_view(self.tableView.file)
                ret, unmatch_cnt, p = rpy_info_dic[self.tableView.file]
                if len(ret) != 0:
                    percentage = unmatch_cnt / len(ret) * 100
                else:
                    percentage = 0

                try:
                    if item is not None and item.index() == index:
                        self.selectTableView.model.item(index.row(), 2).setText(f'{percentage:.2f}%')
                        self.tableView.index = self.selectTableView.model.item(index.row(), 2).index()
                    else:
                        self.treeView.model.data(index)
                except Exception as e:
                    msg = traceback.format_exc()
                    log_print(msg)
                log_print('end saving file!')
                self.show()
                self.raise_()
                self.setEnabled(True)
            if translated_thread is not None and translated_dic is not None:
                translated_thread.join()
                trans_dic = translated_dic
                local_glossary = self.tableView.local_glossary
                for row in self.tableView.selected_rows:
                    if self.tableView.is_original:
                        target = self.tableView.model.item(row, 2).text()
                    else:
                        target = self.tableView.model.item(row, 3).text()
                    if len(target) == 0:
                        continue
                    if local_glossary is not None:
                        for original, replace in local_glossary.items():
                            target = target.replace(original, replace)

                    line_index = int(self.tableView.model.item(row, 0).text())
                    translated = None
                    if self.treeView.is_replace_special_symbols:
                        d = EncodeBrackets(target)
                        translated = get_translated(trans_dic, d)
                    else:
                        if target in translated_dic:
                            translated = trans_dic[target]
                    if translated is None:
                        d = EncodeBrackets(target)
                        encoded = d['encoded'].strip('"')
                        translated = ''
                        if encoded in trans_dic:
                            translated = trans_dic[encoded]
                        log_print(
                            f'{self.tableView.file} Error in line:{str(line_index)}\n{target}\n{encoded}\n{translated}\nError')
                        self.tableView.model.item(row, 4).setText(target)
                        self.tableView.model.item(row, 4).setToolTip(target)
                    else:
                        self.tableView.model.item(row, 4).setText(translated)
                        self.tableView.model.item(row, 4).setToolTip(translated)
                # self.parent.hide()
                if self.autoCopyCheckBox.isChecked():
                    self.tableView.copy_translated_to_cur()
                log_print('translated over')
                translated_dic = None
                translated_thread = None
                self.setEnabled(True)
                self.show()
                self.raise_()
        except Exception:
            msg = traceback.format_exc()
            log_print(msg)
            rpy_lock.release()

    class UpdateThread(QThread):
        update_date = Signal()

        def __init__(self):
            super().__init__()

        def __del__(self):
            self.wait()

        def run(self):
            self.update_date.emit()


class getRpyInfoThread(threading.Thread):
    def __init__(self, p, is_open_filter):
        threading.Thread.__init__(self)
        self.p = p
        self.is_open_filter = is_open_filter

    def run(self):
        try:
            get_rpy_info_from_dir(self.p, self.is_open_filter)
        except Exception as e:
            msg = traceback.format_exc()
            log_print(msg)
            if os.path.isfile('geting_rpy_info'):
                os.remove('geting_rpy_info')


def get_rpy_info_from_dir(select_one, is_open_filter):
    try:
        global rpy_info_dic
        cpu_num = multiprocessing.cpu_count()
        pool = multiprocessing.Pool(cpu_num)
        paths = os.walk(select_one, topdown=False)
        select_one = select_one.replace('\\', '/')
        jobs = []
        for path, dir_lst, file_lst in paths:
            for file_name in file_lst:
                if is_open_filter and not file_name.endswith("rpy"):
                    continue
                i = os.path.join(path, file_name)
                jobs.append(i)

        res = pool.map(get_rpy_info, jobs)
        pool.close()
        pool.join()
        for ret, unmatch_cnt, p in res:
            p = p.replace('\\', '/')
            rpy_info_dic[p] = ret, unmatch_cnt, p
        rpy_lock.acquire()
        f = io.open('rpy_info_got', 'w', encoding='utf-8')
        f.write(select_one)
        f.close()
        rpy_lock.release()
    except:
        msg = traceback.format_exc()
        log_print(msg)
